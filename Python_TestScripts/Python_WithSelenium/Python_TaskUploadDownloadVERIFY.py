import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.webdriver import WebDriver
from selenium.webdriver.ie.service import Service
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

# Adding path of File and load it successfully
File_path = "C:/Users/Parveen/PythonProject_Scratch/FileUpload_Testing.xlsx"
Excel_sheet  = openpyxl.load_workbook(File_path)
sheet_value = Excel_sheet.active

# Now pick up the records
row0 = sheet_value.cell(row=5, column=4)
print(row0.value)

# Updating the values
row1 = sheet_value.cell(row=5, column=4)
row1.value = 999
print(row1.value)

# Save the file with updated records
Excel_sheet.save("C:/Users/Parveen/PythonProject_Scratch/FileUpload_Testing.xlsx")

Dict_list = {}

for i in range( 1, sheet_value.max_row + 1 ):
    if sheet_value.cell(row=i, column=2).value == "Banana": # one by one check the rows
        for j in range ( 1, sheet_value.max_column + 1):    # Pick Up Banana row, count the only 1 column where banana is present.
            print(sheet_value.cell(row=i, column=j).value)
            Dict_list[sheet_value.cell(row=1, column=j).value] = sheet_value.cell(row=i, column=j).value

print(Dict_list)

time.sleep(5)