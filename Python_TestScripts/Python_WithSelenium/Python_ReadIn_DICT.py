import openpyxl

sheet = openpyxl.load_workbook(r"C:\Users\Parveen\PythonProject_Scratch\MOCK_DATA.xlsx")
excel_book = sheet.active

Dict = {}       # Add ros and column in dictionary

excel_book.cell(row=100, column=5).value = "Male"
record_value = excel_book.cell(row=20, column=3).value
print(record_value)

for i in range(1, excel_book.max_row + 1 ): # For reading the rows
    if excel_book.cell(row=i, column=2).value == "Parveen":
        for j in range (1, excel_book.max_column + 1):      # For reading the columns
            print(excel_book.cell(row=i, column=j).value)
            Dict[excel_book.cell(row=1 , column=j).value] = excel_book.cell(row=i, column=j).value
print(Dict)

print("Rows and Column : Read and Write successfully.")

# But the issue in dict is, I'm getting only 1 result instead of all 3,
# because your Dict is being overwritten every time inside the loop.

# So even if "Parveen" appears 3 times, the dictionary is updated again and again
# So write one more code to print 2 records