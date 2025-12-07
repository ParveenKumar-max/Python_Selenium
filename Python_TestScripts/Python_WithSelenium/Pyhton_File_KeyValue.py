import openpyxl

sheet = openpyxl.load_workbook(r"C:\Users\Parveen\PythonProject_Scratch\MOCK_DATA.xlsx")
excel_book = sheet.active

result = []   # list all Parveen rows


#excel_book.cell(row=100, column=5).value = "Male"
#record_value = excel_book.cell(row=20, column=3).value
#print(record_value)

for i in range(1, excel_book.max_row + 1 ): # For reading the rows
    if excel_book.cell(row=i, column=2).value == "Parveen":

        Dict = {}  # add rows and columns in kay value pair

        for j in range (1, excel_book.max_column + 1):      # For reading the columns
            key_pair = excel_book.cell(row=1, column=j).value
            value_pair = excel_book.cell(row=i, column=j).value
            Dict[key_pair] = value_pair
        result.append(Dict)
        #print(excel_book.cell(row=i, column=j).value)
        #Dict[excel_book.cell(row=1 , column=j).value] = excel_book.cell(row=i, column=j).value
        print(Dict) # it will print the value in Dictionary format
print(result)       # it will print the value in List Dictionary format

print("Rows and Column : Read and Write successfully.")

# Now the output is as expected, three key value pair is showing in the output
