import openpyxl
#Pytest framework, you need the openpyxl library primarily for data-driven testing and test reporting,
# .whenever your test data or output needs to be stored in or retrieved from Excel spreadsheet files (.xlsx).

sheet = openpyxl.load_workbook(r"C:\Users\Parveen\PythonProject_Scratch\MOCK_DATA.xlsx")
excel_book = sheet.active

record_value = excel_book.cell(row=20, column=3).value
print(record_value)

for i in range(1, excel_book.max_row + 1 ): # For reading the rows
    if excel_book.cell(row=i, column=2).value == "Parveen":
        for j in range (2, excel_book.max_column + 1):      # For reading the columns
            print(excel_book.cell(row=i, column=j).value)

