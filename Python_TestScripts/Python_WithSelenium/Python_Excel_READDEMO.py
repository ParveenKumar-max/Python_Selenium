import openpyxl

Book = openpyxl.load_workbook(r"C:\Users\Parveen\PythonProject_Scratch\MOCK_DATA.xlsx")
sheet = Book.active
print(sheet.max_row)
print(sheet.max_row + 1)

cell= sheet.cell(row=1,column=4)
print(cell.value)

sheet.cell (  row = 10, column = 3 ).value = "Chaudhary"
print(sheet.cell (  row = 10, column = 2 ).value)
print(sheet['C11'].value)
Book.save(r"C:\Users\Parveen\PythonProject_Scratch\MOCK_DATA.xlsx")