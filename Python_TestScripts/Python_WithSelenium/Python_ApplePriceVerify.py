import time

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.webdriver import WebDriver
from selenium.webdriver.ie.service import Service
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

# Adding path of File and load it successfully
File_path = "C:/Users/Parveen/PythonProject_Scratch/FileUpload_Testing.xlsx"
Excel_sheet  = openpyxl.load_workbook(File_path)
sheet_value = Excel_sheet.active

# Now pick up the records
row0 = sheet_value.cell(row=3, column=4)
#print(row0.value)

# Updating the value of APPLE
row1 = sheet_value.cell(row=3, column=4)
row1.value = row1.value + 5
#print(row1.value)

# Save the file with updated records
Excel_sheet.save("C:/Users/Parveen/PythonProject_Scratch/FileUpload_Testing.xlsx")

#Create 1 empty variable
Apple_price  =  None

#Pick Up the records on the basis of APPLE
for x in range(1, sheet_value.max_row + 1):   # Count the rows one by one
    if sheet_value.cell(row=x, column=2).value == "Apple":
        #for y in range (1, sheet_value.max_column + 1):
            Apple_price = sheet_value.cell(row=x, column=4) #Pick the price of APPLE
            print(f"Excel file Apple Value:. {Apple_price.value}")


#Define Folder where we added the msg driver
Folder_path = "C:/Users/Default/Documents/msedgedriver.exe"
Service_url = Service(Folder_path)
driver = webdriver.Edge(service=Service_url)

#Open the URL
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")

#Add wait condition for loading the DOM
wait = WebDriverWait(driver, 10)
wait.until(expected_conditions.visibility_of_element_located((By.XPATH,"//div[@class = 'sc-fLseNd igTRoD']")))

#ADD File Path and Upload the file
File_path = "C:/Users/Parveen/PythonProject_Scratch/FileUpload_Testing.xlsx"
File_upload = driver.find_element(By.XPATH, "//input[@type='file']")

#Add send keys to upload the File
File_upload.send_keys(File_path)
print("File Uploaded Successfully")

wait.until(expected_conditions.visibility_of_element_located((By.XPATH,"//div[@id='row-1']//div[@id = 'cell-4-undefined']")))
Text_apple_value = int(driver.find_element(By.XPATH,"//div[@id='row-1']//div[@id = 'cell-4-undefined']").text)

print(f"Get Apple value from GUI: {Text_apple_value}")

assert Text_apple_value == Apple_price.value,  f"Mismatch ! Excel ={Apple_price}, GUI={Text_apple_value}"

print("Apple price is matching")


time.sleep(5)